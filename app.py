import math
import tempfile
from io import BytesIO
from datetime import datetime

import numpy as np
import pandas as pd
import plotly.graph_objects as go
import streamlit as st
from openpyxl import Workbook

import volatility_fitting_daily as vf


SPX_DIV_YIELD = 0.0134
SPY_DIVS = [(0.25, 1.90), (0.50, 2.10), (0.75, 1.90), (1.00, 1.92)]
DATA_FETCH_VERSION = "six-tenors-no-1w-diagnostics"


st.set_page_config(
    page_title="Volatility Surface Dashboard",
    page_icon="",
    layout="wide",
)


st.markdown(
    """
    <style>
    .block-container { padding-top: 1.4rem; }
    h1, h2, h3 { letter-spacing: 0; }
    div[data-testid="stMetric"] {
        border: 1px solid #e7e7e7;
        border-radius: 8px;
        padding: 12px 14px;
        background: #ffffff;
    }
    div[data-testid="stMetricValue"] {
        font-size: 1.05rem;
        white-space: normal;
        overflow-wrap: anywhere;
    }
    </style>
    """,
    unsafe_allow_html=True,
)


def get_secret(name):
    try:
        return st.secrets.get(name, "")
    except Exception:
        return ""


def load_excel_from_upload(uploaded_file):
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(uploaded_file.getvalue())
        tmp_path = tmp.name
    return vf.load_xlsx(tmp_path)


def sample_workbook_bytes():
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "SPX"
    ws["B1"] = "Spot"
    ws["C1"] = 5667.56
    ws["A2"] = "Expiry date"
    ws["B2"] = "2025-03-21"
    ws["P2"] = 1
    ws["Q2"] = "m"
    ws["R2"] = 4.36
    ws["D3"] = "SPX tenor rate"
    ws["H3"] = "SPX tenor rate"
    ws["L3"] = "SPX tenor rate"
    ws["B4"] = "Expiry 1"
    ws["C4"] = "Tenor years"
    ws["F4"] = "Expiry 2"
    ws["G4"] = "Tenor years"
    ws["J4"] = "Expiry 3"
    ws["K4"] = "Tenor years"
    ws["A5"] = "Strike"
    ws["B5"] = "C"
    ws["C5"] = "P"
    ws["E5"] = "Strike"
    ws["F5"] = "C"
    ws["G5"] = "P"
    ws["I5"] = "Strike"
    ws["J5"] = "C"
    ws["K5"] = "P"
    ws["A23"] = "SPY"
    ws["B23"] = "Spot"
    ws["C23"] = 566.76
    ws["A24"] = "SPY tenor rate"
    ws["D24"] = 0.045
    ws["H24"] = 0.045
    ws["L24"] = 0.0435
    ws["B25"] = "Expiry 1"
    ws["C25"] = "Tenor years"
    ws["F25"] = "Expiry 2"
    ws["G25"] = "Tenor years"
    ws["J25"] = "Expiry 3"
    ws["K25"] = "Tenor years"
    ws["A26"] = "Strike"
    ws["B26"] = "C"
    ws["C26"] = "P"
    ws["E26"] = "Strike"
    ws["F26"] = "C"
    ws["G26"] = "P"
    ws["I26"] = "Strike"
    ws["J26"] = "C"
    ws["K26"] = "P"
    ws["A43"] = "This template shows the required cell layout. Replace placeholders with actual quotes."
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue()


@st.cache_data(show_spinner=False)
def load_latest_cached(fred_api_key, version):
    return vf.load_latest_data(fred_api_key)


@st.cache_data(show_spinner=False)
def compute_asset_data(label, spot, divs, ydiv, quotes, curve, american, per_tenor):
    rate_curve = vf.build_curve(per_tenor, curve)
    div_list = [] if ydiv > 0 else divs

    if american:
        repo = vf.fit_repo_am(spot, rate_curve, div_list, quotes)
    else:
        repo = vf.fit_repo_eu(spot, rate_curve, div_list, quotes)

    ivs = vf.compute_ivs(spot, rate_curve, div_list, repo, quotes, american)
    repo_rows = []
    for t in sorted(repo):
        r = vf.interp_rate(rate_curve, t)
        s_eff = spot - vf.pv_divs(div_list, t, rate_curve)
        forward = s_eff * math.exp((r - repo[t]) * t)
        repo_rows.append({
            "asset": label,
            "tenor": t,
            "rate": r,
            "repo": repo[t],
            "forward": forward,
        })

    iv_rows = []
    for t, rows in ivs.items():
        r = vf.interp_rate(rate_curve, t)
        forward = (spot - vf.pv_divs(div_list, t, rate_curve)) * math.exp((r - repo[t]) * t)
        for strike, is_call, iv in rows:
            iv_rows.append({
                "asset": label,
                "tenor": t,
                "strike": strike,
                "option": "Call" if is_call else "Put",
                "iv": iv,
                "iv_percent": 100 * iv,
                "log_moneyness": math.log(strike / forward),
                "forward": forward,
            })

    return pd.DataFrame(repo_rows), pd.DataFrame(iv_rows)


def fit_display_surface(iv_df, moneyness_limit):
    filtered = iv_df[iv_df["log_moneyness"].abs() <= moneyness_limit].copy()
    coef_rows = []

    for tenor in sorted(filtered["tenor"].unique()):
        part = filtered[filtered["tenor"] == tenor].copy()
        otm = part[((part["option"] == "Call") & (part["log_moneyness"] > 0))
                   | ((part["option"] == "Put") & (part["log_moneyness"] < 0))]
        fit_part = otm if len(otm) >= 3 else part
        if len(fit_part) < 3:
            continue

        x = fit_part["log_moneyness"].to_numpy()
        y = fit_part["iv"].to_numpy()
        a, b, c = vf.fit_quad(x, y)
        y_hat = a + b * x + c * x * x
        resid = y - y_hat
        rmse = float(np.sqrt(np.mean(resid * resid)))
        denom = float(np.sum((y - np.mean(y)) ** 2))
        r2 = float(1 - np.sum(resid * resid) / denom) if denom > 0 else np.nan

        coef_rows.append({
            "asset": fit_part["asset"].iloc[0],
            "tenor": tenor,
            "a": a,
            "b": b,
            "c": c,
            "n_fit": len(fit_part),
            "rmse": rmse,
            "rmse_pct": 100 * rmse,
            "r2": r2,
            "butterfly_ok": c >= -1e-8,
        })

    coef_df = pd.DataFrame(coef_rows)
    if coef_df.empty:
        return filtered, coef_df
    return filtered, add_calendar_flags(filtered, coef_df)


def add_calendar_flags(iv_df, coef_df):
    out = coef_df.sort_values("tenor").copy()
    out["calendar_ok"] = True
    out["calendar_violations"] = 0

    tenors = list(out["tenor"])
    for i in range(1, len(tenors)):
        prev_t, curr_t = tenors[i - 1], tenors[i]
        prev_iv = iv_df[iv_df["tenor"] == prev_t]
        curr_iv = iv_df[iv_df["tenor"] == curr_t]
        lo = max(float(prev_iv["log_moneyness"].min()), float(curr_iv["log_moneyness"].min()))
        hi = min(float(prev_iv["log_moneyness"].max()), float(curr_iv["log_moneyness"].max()))
        if lo >= hi:
            continue

        x = np.linspace(lo, hi, 40)
        p = out[out["tenor"] == prev_t].iloc[0]
        q = out[out["tenor"] == curr_t].iloc[0]
        prev_vol = p["a"] + p["b"] * x + p["c"] * x * x
        curr_vol = q["a"] + q["b"] * x + q["c"] * x * x
        prev_total_var = prev_vol * prev_vol * prev_t
        curr_total_var = curr_vol * curr_vol * curr_t
        violations = int(np.sum(curr_total_var + 1e-6 < prev_total_var))
        out.loc[out["tenor"] == curr_t, "calendar_violations"] = violations
        out.loc[out["tenor"] == curr_t, "calendar_ok"] = violations == 0

    return out


def fitted_surface_figure(asset, iv_df, coef_df):
    fig = go.Figure()
    if iv_df.empty or coef_df.empty:
        return fig

    tenor_ranges = {}
    for tenor in sorted(coef_df["tenor"].unique()):
        part = iv_df[iv_df["tenor"] == tenor]
        if not part.empty:
            tenor_ranges[tenor] = (float(part["log_moneyness"].min()), float(part["log_moneyness"].max()))

    if not tenor_ranges:
        return fig

    # Use a broad display range, but mask each tenor outside its own observed
    # range. That avoids quadratic extrapolation spikes without collapsing the
    # surface into a skinny ribbon when one tenor has a narrower strike range.
    x_min = max(float(iv_df["log_moneyness"].quantile(0.03)), -0.08)
    x_max = min(float(iv_df["log_moneyness"].quantile(0.97)), 0.06)
    x_grid = np.linspace(x_min, x_max, 60)
    tenors = sorted(coef_df["tenor"].unique())

    z_rows = []
    for tenor in tenors:
        row = coef_df.loc[coef_df["tenor"] == tenor].iloc[0]
        fitted = 100 * (row["a"] + row["b"] * x_grid + row["c"] * x_grid * x_grid)
        lo, hi = tenor_ranges[tenor]
        fitted = np.where((x_grid >= lo) & (x_grid <= hi), fitted, np.nan)
        z_rows.append(np.clip(fitted, 5, 35))

    fig.add_trace(go.Surface(
        x=x_grid,
        y=tenors,
        z=np.array(z_rows),
        colorscale="Viridis",
        opacity=0.86,
        name="Fitted surface",
        connectgaps=False,
        colorbar={"title": "IV %"},
    ))

    fig.add_trace(go.Scatter3d(
        x=iv_df["log_moneyness"],
        y=iv_df["tenor"],
        z=iv_df["iv_percent"],
        mode="markers",
        marker={"size": 4, "color": iv_df["iv_percent"], "colorscale": "Turbo"},
        name="Market IV points",
    ))

    fig.update_layout(
        title=f"{asset} Implied Volatility Surface",
        height=620,
        margin={"l": 0, "r": 0, "t": 50, "b": 0},
        scene={
            "xaxis_title": "log(K / F)",
            "yaxis_title": "Time to expiry",
            "zaxis_title": "IV %",
        },
    )
    return fig


def smile_figure(asset, iv_df):
    fig = go.Figure()
    for tenor in sorted(iv_df["tenor"].unique()):
        part = iv_df[iv_df["tenor"] == tenor].sort_values("log_moneyness")
        for option, symbol in [("Put", "circle"), ("Call", "diamond")]:
            side = part[part["option"] == option]
            fig.add_trace(go.Scatter(
                x=side["log_moneyness"],
                y=side["iv_percent"],
                mode="markers",
                marker={"symbol": symbol, "size": 7},
                name=f"{option} t={tenor:.3f}",
                text=option + " K=" + side["strike"].round(2).astype(str),
                legendgroup=f"{tenor:.3f}",
                showlegend=tenor == sorted(iv_df["tenor"].unique())[0],
            ))

        fit_part = part[((part["option"] == "Call") & (part["log_moneyness"] > 0))
                        | ((part["option"] == "Put") & (part["log_moneyness"] < 0))]
        if len(fit_part) >= 3:
            a, b, c = vf.fit_quad(fit_part["log_moneyness"].to_numpy(), fit_part["iv"].to_numpy())
            x = np.linspace(float(part["log_moneyness"].min()), float(part["log_moneyness"].max()), 80)
            y = 100 * (a + b * x + c * x * x)
            fig.add_trace(go.Scatter(
                x=x,
                y=y,
                mode="lines",
                line={"width": 2},
                name=f"Fit t={tenor:.3f}",
                legendgroup=f"{tenor:.3f}",
                showlegend=True,
            ))

    fig.update_layout(
        title=f"{asset} Volatility Smiles",
        height=420,
        margin={"l": 0, "r": 0, "t": 50, "b": 0},
        xaxis_title="log(K / F)",
        yaxis_title="IV %",
        legend_title="Tenor",
    )
    return fig


def format_percent_table(df, cols):
    out = df.copy()
    for col in cols:
        if col in out:
            out[col] = 100 * out[col]
    return out


def format_snapshot_time(value):
    if not value:
        return datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    try:
        return datetime.fromisoformat(value).astimezone().strftime("%Y-%m-%d %H:%M:%S %Z")
    except Exception:
        return str(value)


def quality_summary(coef_df):
    if coef_df.empty:
        return "No fit", "No fit", "No fit"
    avg_rmse = coef_df["rmse_pct"].mean()
    avg_r2 = coef_df["r2"].dropna().mean()
    arb_ok = bool(coef_df["butterfly_ok"].all() and coef_df["calendar_ok"].all())
    return f"{avg_rmse:.3f}%", f"{avg_r2:.3f}", "Pass" if arb_ok else "Review"


st.title("Volatility Surface Dashboard")
st.caption("SPX/SPY implied volatility construction from Excel inputs or latest market data.")

with st.sidebar:
    st.header("Inputs")
    mode = st.radio("Data source", ["Fetch latest data", "Upload Excel file"])
    asset_choice = st.radio("Asset", ["SPX", "SPY", "Both"], horizontal=True)
    moneyness_limit = st.slider(
        "Moneyness range |log(K/F)|",
        min_value=0.01,
        max_value=0.12,
        value=0.06,
        step=0.01,
        help="Controls which IV points are used in the displayed fit and charts.",
    )

    fred_key = ""
    uploaded = None
    if mode == "Fetch latest data":
        default_key = get_secret("FRED_API_KEY")
        fred_key = st.text_input("FRED API key", value=default_key, type="password")
        st.caption("For hosting, add this as FRED_API_KEY in Streamlit secrets.")
    else:
        uploaded = st.file_uploader("Upload OptionData.xlsx", type=["xlsx"])
        st.download_button(
            "Download Excel layout template",
            sample_workbook_bytes(),
            file_name="OptionData_template.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
        st.caption("Upload mode expects the original project workbook cell layout.")

    run_clicked = st.button("Run analysis", type="primary", use_container_width=True)

st.markdown(
    "This dashboard fits a repo curve, computes implied volatilities, filters option data by moneyness, "
    "and visualizes the fitted quadratic surface."
)

if not run_clicked:
    st.info("Choose a data source in the sidebar and click Run analysis.")
    st.stop()

try:
    with st.spinner("Loading market data..."):
        if mode == "Fetch latest data":
            if not fred_key:
                st.error("Enter a FRED API key or add FRED_API_KEY to Streamlit secrets.")
                st.stop()
            data = load_latest_cached(fred_key, DATA_FETCH_VERSION)
            source_label = "Latest available yfinance option chains + FRED Treasury curve"
        else:
            if uploaded is None:
                st.error("Upload an OptionData.xlsx file first.")
                st.stop()
            data = load_excel_from_upload(uploaded)
            source_label = uploaded.name
except Exception as exc:
    st.error(str(exc))
    st.stop()

assets = []
if asset_choice in ("SPX", "Both"):
    assets.append(("SPX", data["spx_spot"], [], SPX_DIV_YIELD, data["spx_quotes"], False, data["spx_rates"]))
if asset_choice in ("SPY", "Both"):
    assets.append(("SPY", data["spy_spot"], SPY_DIVS, 0.0, data["spy_quotes"], True, data["spy_rates"]))

st.subheader("Run Summary")
cols = st.columns(4)
cols[0].metric("Source", mode)
cols[1].metric("SPX spot", f"{data['spx_spot']:,.2f}")
cols[2].metric("SPY spot", f"{data['spy_spot']:,.2f}")
cols[3].metric("Snapshot time", format_snapshot_time(data.get("fetched_at")))
st.caption(source_label)

for label, spot, divs, ydiv, quotes, american, rates in assets:
    with st.spinner(f"Computing {label} surface..."):
        repo_df, raw_iv_df = compute_asset_data(label, spot, divs, ydiv, quotes, data["curve"], american, rates)
        iv_df, coef_df = fit_display_surface(raw_iv_df, moneyness_limit)

    st.divider()
    st.header(label)

    avg_rmse, avg_r2, arb_status = quality_summary(coef_df)
    summary_cols = st.columns(7)
    summary_cols[0].metric("Spot", f"{spot:,.2f}")
    summary_cols[1].metric("Model", "American" if american else "European")
    summary_cols[2].metric("Tenors", len(repo_df))
    summary_cols[3].metric("IV points", f"{len(iv_df)} / {len(raw_iv_df)}")
    summary_cols[4].metric("Avg RMSE", avg_rmse)
    summary_cols[5].metric("Avg R2", avg_r2)
    summary_cols[6].metric("Arbitrage flags", arb_status)

    if coef_df.empty:
        st.warning("Not enough IV points inside the selected moneyness range to fit this asset.")
        continue
    if arb_status == "Pass":
        st.success("Simple no-arbitrage screen passed: nonnegative smile curvature and nondecreasing fitted total variance across adjacent tenors.")
    else:
        st.warning("Review no-arbitrage flags: at least one tenor has negative smile curvature or a calendar total-variance decrease.")

    tab_surface, tab_smiles, tab_diagnostics, tab_tables = st.tabs(["Surface", "Smiles", "Diagnostics", "Tables"])

    with tab_surface:
        st.plotly_chart(fitted_surface_figure(label, iv_df, coef_df), use_container_width=True, key=f"{label}-surface")
        st.caption("Fitted surface from quadratic smile fits by expiry. Markers show option-implied volatility points; the surface is only plotted over observed moneyness ranges to avoid extrapolation artifacts.")

    with tab_smiles:
        st.plotly_chart(smile_figure(label, iv_df), use_container_width=True, key=f"{label}-smiles")

    with tab_diagnostics:
        diagnostics = coef_df[[
            "tenor", "n_fit", "rmse_pct", "r2", "butterfly_ok", "calendar_ok", "calendar_violations"
        ]].copy()
        diagnostics["butterfly"] = diagnostics["butterfly_ok"].map({True: "Pass", False: "Review"})
        diagnostics["calendar"] = diagnostics["calendar_ok"].map({True: "Pass", False: "Review"})
        diagnostics = diagnostics.drop(columns=["butterfly_ok", "calendar_ok"])
        st.dataframe(diagnostics, use_container_width=True, hide_index=True)
        st.caption("Butterfly flag uses quadratic smile convexity. Calendar flag checks fitted total variance across adjacent tenors over their overlapping observed moneyness range.")

    with tab_tables:
        table_cols = st.columns(3)
        with table_cols[0]:
            st.markdown("**Repo Curve**")
            repo_show = format_percent_table(repo_df, ["rate", "repo"])
            st.dataframe(repo_show, use_container_width=True, hide_index=True)
        with table_cols[1]:
            st.markdown("**Surface Fit**")
            st.dataframe(coef_df, use_container_width=True, hide_index=True)
        with table_cols[2]:
            st.markdown("**Filtered Implied Vols**")
            st.dataframe(iv_df, use_container_width=True, hide_index=True)

    csv = iv_df.to_csv(index=False).encode("utf-8")
    st.download_button(
        f"Download {label} IV data",
        csv,
        file_name=f"{label.lower()}_implied_vols.csv",
        mime="text/csv",
    )
